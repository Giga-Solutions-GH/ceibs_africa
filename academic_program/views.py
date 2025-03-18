import csv
import datetime
import tempfile
from io import BytesIO

import openpyxl
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.db.models.functions import ExtractYear
from django.forms import modelformset_factory
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.views.decorators.http import require_POST
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.workbook import Workbook

from academic_program import models, forms
from academic_program.forms import AttendanceForm, ProgramScheduleForm, StudentForm
from academic_program.models import Program, Course, CourseParticipant, Lecturer, ProgramParticipant, Attendance, \
    ProgramSchedule, ProgramScheduleDate, ProgramCover, ProgramType
from account.models import CustomUser, Role
from finance.models import ProgramFees, FinanceStatement, StudentFinance
from students.models import StudentDetail, StudentEnrollment
from .tasks import send_student_enrollment_email, send_program_participant_welcome_email, process_excel_import_task, \
    convert_selected_students_to_alumni_task


# Create your views here.
def program_type_overview(request):
    program_type = models.ProgramType.objects.all()
    context = {
        'program_types': program_type
    }
    return render(request, "academic_program/layouts/program_and_courses/program_types.html", context=context)


def programs(request, program_type):
    # Get the ProgramType object (or 404)
    program_type_obj = get_object_or_404(ProgramType, name=program_type)

    # Base queryset: all programs of this type that are active (flag True)
    qs = Program.objects.filter(program_type=program_type_obj, flag=True)

    # Filter by year (using the program start_date year)
    year = request.GET.get('year')
    if year:
        qs = qs.filter(start_date__year=year)

    # Filter by search query on program name
    search_query = request.GET.get('search')
    if search_query:
        qs = qs.filter(program_name__icontains=search_query)

    # Filter by program cover (if provided, using case‐insensitive matching)
    cover = request.GET.get('cover')
    if cover:
        qs = qs.filter(program_cover__name__icontains=cover)

    # Group programs into active and ended
    active_programs = qs.filter(program_ended=False).order_by('start_date')
    ended_programs = qs.filter(program_ended=True).order_by('start_date')

    # For the filter dropdowns
    all_years = sorted({str(p.start_date.year) for p in qs if p.start_date})
    all_covers = ProgramCover.objects.all()

    context = {
        'program_type': program_type_obj.name,
        'active_programs': active_programs,
        'ended_programs': ended_programs,
        'year_filter': year,
        'search_query': search_query,
        'cover_filter': cover,
        'all_years': all_years,
        'all_covers': all_covers,
    }
    return render(request, "academic_program/layouts/program_and_courses/programs.html", context)


def program_statistics(request):
    # Number of programs by type
    programs_by_type = Program.objects.values('program_type__name').annotate(count=Count('id')).order_by(
        'program_type__name')

    # Number of courses per program
    courses_per_program = Course.objects.values('program__program_name').annotate(count=Count('id')).order_by(
        'program__program_name')

    # Number of students per program
    students_per_program = StudentEnrollment.objects.values('program__program_name').annotate(
        count=Count('id')).order_by('program__program_name')

    # Number of students per course
    students_per_course = Course.objects.annotate(student_count=Count('participants')).values('course_name',
                                                                                              'student_count').order_by(
        'course_name')

    # Aggregate number of students per program cover and year
    program_cover_stats = (
        Program.objects
        .values('program_cover__name', 'program_cover')
        .annotate(
            year=ExtractYear('start_date'),
            enrolled_students=Count('studentenrollment')
        )
        .order_by('program_cover', 'year')
    )

    # Prepare data for each program cover
    program_cover_data = {}
    for entry in program_cover_stats:
        cover_name = entry['program_cover__name']
        year = entry['year']
        count = entry['enrolled_students']

        if cover_name not in program_cover_data:
            program_cover_data[cover_name] = {}

        if year not in program_cover_data[cover_name]:
            program_cover_data[cover_name][year] = 0

        program_cover_data[cover_name][year] += count

    context = {
        'programs_by_type': programs_by_type,
        'courses_per_program': courses_per_program,
        'students_per_program': students_per_program,
        'students_per_course': students_per_course,
        'program_cover_data': program_cover_data,
    }

    return render(request, 'academic_program/layouts/program_and_courses/program_statistics.html', context)


def export_program_statistics(request):
    # Prepare the HTTP response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="program_statistics.csv"'

    # Create a CSV writer
    writer = csv.writer(response)

    # Write the header for Programs by Type
    writer.writerow(['Programs by Type'])
    writer.writerow(['Program Type', 'Count'])

    # Fetch data
    programs_by_type = Program.objects.values('program_type__name').annotate(count=Count('id')).order_by(
        'program_type__name')
    for item in programs_by_type:
        writer.writerow([item['program_type__name'], item['count']])

    writer.writerow([])  # Empty row for separation

    # Write the header for Courses per Program
    writer.writerow(['Courses per Program'])
    writer.writerow(['Program Name', 'Number of Courses'])

    # Fetch data
    courses_per_program = Course.objects.values('program__program_name').annotate(count=Count('id')).order_by(
        'program__program_name')
    for item in courses_per_program:
        writer.writerow([item['program__program_name'], item['count']])

    writer.writerow([])  # Empty row for separation

    # Write the header for Students per Program
    writer.writerow(['Students per Program'])
    writer.writerow(['Program Name', 'Number of Students'])

    # Fetch data
    students_per_program = StudentEnrollment.objects.values('program__program_name').annotate(
        count=Count('id')).order_by('program__program_name')
    for item in students_per_program:
        writer.writerow([item['program__program_name'], item['count']])

    writer.writerow([])  # Empty row for separation

    # Write the header for Students per Course
    writer.writerow(['Students per Course'])
    writer.writerow(['Course Name', 'Number of Students'])

    # Fetch data
    students_per_course = Course.objects.annotate(student_count=Count('participants')).values('course_name',
                                                                                              'student_count').order_by(
        'course_name')
    for item in students_per_course:
        writer.writerow([item['course_name'], item['student_count']])

    writer.writerow([])  # Empty row for separation

    # Write the header for Program Cover Statistics
    writer.writerow(['Program Cover Statistics'])
    writer.writerow(['Program Cover', 'Year', 'Number of Enrolled Students'])

    # Fetch data
    program_cover_stats = (
        Program.objects
        .values('program_cover__name', 'program_cover')
        .annotate(
            year=ExtractYear('start_date'),
            enrolled_students=Count('studentenrollment')
        )
        .order_by('program_cover', 'year')
    )

    # Prepare data for each program cover
    program_cover_data = {}
    for entry in program_cover_stats:
        cover_name = entry['program_cover__name']
        year = entry['year']
        count = entry['enrolled_students']

        if cover_name not in program_cover_data:
            program_cover_data[cover_name] = {}

        if year not in program_cover_data[cover_name]:
            program_cover_data[cover_name][year] = 0

        program_cover_data[cover_name][year] += count

    # Write program cover data
    for cover_name, years in program_cover_data.items():
        for year, count in years.items():
            writer.writerow([cover_name, year, count])

    return response


def program_module_overview(request):
    """
    Groups all programs by their ProgramCover and sends them to the template
    along with a list of distinct covers for client-side filtering.
    """
    # Fetch all programs (adjust if you only want flagged or active ones)
    programs = Program.objects.all().select_related('program_cover').order_by('program_cover__name', 'program_name')

    # Group the programs by cover
    cover_dict = {}
    for prog in programs:
        cover = prog.program_cover  # could be None if not set
        if cover not in cover_dict:
            cover_dict[cover] = []
        cover_dict[cover].append(prog)

    # Build a list of covers actually used (excluding None). We can filter by these
    used_covers = []
    for cover in cover_dict.keys():
        if cover:
            used_covers.append(cover)
    # Sort covers by name
    used_covers.sort(key=lambda c: c.name)

    context = {
        'cover_dict': cover_dict,  # { ProgramCover or None : [Program, ...] }
        'used_covers': used_covers,  # for filter dropdown
    }
    return render(request, 'academic_program/layouts/program_and_courses/program_module_overview.html', context)


def add_student_from_program(request, program_id):
    program = get_object_or_404(Program, id=program_id)
    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            # Create student detail
            student = form.save()

            # Create associated user (using unique_id as temporary password; change as needed)
            user = CustomUser.objects.create_user(
                username=student.unique_id,
                email=student.email,
                password=student.unique_id,
                first_name=student.first_name,
                last_name=student.last_name,
                is_student=True
            )
            user.save()

            # Link student detail to user
            student.user = user
            student.save()

            # Assign "Student" role (create if not exists)
            student_role, _ = Role.objects.get_or_create(name='Student')
            user.roles.add(student_role)
            user.save()

            # Enroll student in the program
            enrolment = StudentEnrollment.objects.create(
                student=student,
                program=program,
                start_date=program.start_date,
                end_date=program.end_date,
                status="Active"
            )

            # Enroll student in all courses under the program
            courses = Course.objects.filter(program=program)
            for course in courses:
                CourseParticipant.objects.create(
                    course=course,
                    student=student
                )

            program_fees = ProgramFees.objects.get_or_create(program=program)
            finance_statement = FinanceStatement.objects.get_or_create(program_fees=program_fees)
            new_student = enrolment.student
            new_student_finance = StudentFinance.objects.create(
                student=new_student,
                finance_statement=finance_statement,
            )

            # Send enrollment welcome email asynchronously
            send_student_enrollment_email.delay(student.id, program.id)

            messages.success(request,
                             "Student added and enrolled in the program and courses. A welcome email is being sent in the background!")
            return redirect('add_student_from_program_page', program_id=program_id)
        else:
            messages.error(request, "Form is not valid. Please correct the errors.")
    else:
        form = StudentForm()

    context = {
        'form': form,
        'program': program,
    }
    return render(request, 'academic_program/layouts/add_student_from_program_page.html', context)


def add_program_participant(request, program_id):
    program = get_object_or_404(Program, id=program_id)
    if request.method == 'POST':
        form = forms.ProgramParticipantForm(request.POST)
        if form.is_valid():
            participant = form.save(commit=False)
            participant.program = program
            participant.save()
            messages.success(request, 'Participant added to the program successfully!')
            # Send welcome email to participant in the background
            send_program_participant_welcome_email.delay(participant.id, program.id)
            return redirect('add_program_participant', program_id=program.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = forms.ProgramParticipantForm()

    context = {
        'form': form,
        'program': program
    }
    return render(request, 'academic_program/layouts/add_program_participant.html', context)


def program_participant_detail(request, participant_id):
    participant = get_object_or_404(models.ProgramParticipant, id=participant_id)

    if request.method == 'POST':
        form = forms.ProgramParticipantForm(request.POST, instance=participant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Participant details updated successfully!')
            return redirect('program_participant_detail', participant_id=participant.id)
    else:
        form = forms.ProgramParticipantForm(instance=participant)

    context = {
        'form': form,
        'participant': participant,
    }
    return render(request, 'academic_program/layouts/program_and_courses/program_participant_detail.html', context)


def delete_participant(request, participant_id):
    participant = models.ProgramParticipant.objects.get(id=participant_id)
    participant.flag = False
    participant.save()
    return redirect('programs')


def student_list(request, program_name, program_type):
    """Displays list of students or participants for a program,
    with badges if they are the PO or APO.
    Does not handle 'make_po' or 'make_apo' logic here."""
    program = get_object_or_404(Program, program_name=program_name, flag=True)

    if request.method == "POST":
        # Convert selected to alumni
        if 'convert_selected_alumni' in request.POST:
            selected_student_ids = request.POST.getlist('selected_students')
            if selected_student_ids:
                student_id_ints = [int(sid) for sid in selected_student_ids]
                convert_selected_students_to_alumni_task.delay(student_id_ints, program.id)
                messages.success(request, "Selected students are being converted to Alumni in the background!")
            else:
                messages.warning(request, "No students selected for alumnus conversion.")
            return redirect('academic_program:student_list', program_name=program_name, program_type=program_type)

        # Handle file upload for bulk import
        file = request.FILES.get('student_file', None)
        if file:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                for chunk in file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            process_excel_import_task.delay(tmp_path, program.id)
            messages.success(request, "File uploaded successfully. The records are being processed in the background.")
            return redirect('academic_program:student_list', program_name=program_name, program_type=program_type)

    # Retrieve students
    if program.alumni_program:
        active_enrollments = StudentEnrollment.objects.filter(
            program=program, active=True
        ).exclude(status__in=["Alumnus", "Completed", "Inactive", "Deferred"])
        students = StudentDetail.objects.filter(studentenrollment__in=active_enrollments).distinct()
    else:
        students = ProgramParticipant.objects.filter(program=program)

    # Annotate each student with a PO/APO badge
    for student in students:
        badge = ""
        # Because StudentDetail is used for alumni programs,
        # ProgramParticipant is for non-alumni. We'll check if the
        # Program's officer is referencing this student's active enrollment.
        if program.program_officer and program.program_officer.student == student:
            badge += '<span class="badge bg-success me-1">PO</span>'
        if program.assistant_program_officer and program.assistant_program_officer.student == student:
            badge += '<span class="badge bg-warning">APO</span>'
        student.badge = badge

    # Basic stats
    count = students.count() if hasattr(students, 'count') else len(students)
    if program.alumni_program:
        male_count = students.filter(gender='Male').count()
        female_count = students.filter(gender='Female').count()
    else:
        # For ProgramParticipant, we check if there's a gender attribute
        first = students.first()
        if first and hasattr(first, 'gender'):
            male_count = students.filter(gender='Male').count()
            female_count = students.filter(gender='Female').count()
        else:
            male_count = female_count = 0
    active_count = count

    context = {
        'students': students,
        'count': count,
        'male_count': male_count,
        'female_count': female_count,
        'program_name': program_name,
        'program': program,
        'program_type': program.program_type,
        'active_count': active_count,
    }
    return render(request, "academic_program/layouts/student_list.html", context)


def make_program_officer(request, program_name, program_type, student_id):
    """Makes the specified student the Program Officer for the given program."""
    program = get_object_or_404(Program, program_name=program_name, flag=True)

    # The student must exist
    try:
        student_to_update = StudentDetail.objects.get(pk=student_id)
    except StudentDetail.DoesNotExist:
        messages.error(request, "Student not found.")
        return redirect('academic_program:student_list', program_name=program_name, program_type=program_type)

    # The student must have an active enrollment in this program
    enrollment = StudentEnrollment.objects.filter(
        student=student_to_update, program=program, active=True
    ).first()
    if enrollment:
        program.program_officer = enrollment
        program.save()
        messages.success(request,
                         f"{student_to_update.first_name} {student_to_update.last_name} is now the Program Officer.")
    else:
        messages.error(request, "No active enrollment found for this student in this program.")
    return redirect('academic_program:student_list', program_name=program_name, program_type=program_type)


def make_assistant_program_officer(request, program_name, program_type, student_id):
    """Makes the specified student the Assistant Program Officer for the given program."""
    program = get_object_or_404(Program, program_name=program_name, flag=True)

    try:
        student_to_update = StudentDetail.objects.get(pk=student_id)
    except StudentDetail.DoesNotExist:
        messages.error(request, "Student not found.")
        return redirect('academic_program:student_list', program_name=program_name, program_type=program_type)

    enrollment = StudentEnrollment.objects.filter(
        student=student_to_update, program=program, active=True
    ).first()
    if enrollment:
        program.assistant_program_officer = enrollment
        program.save()
        messages.success(request,
                         f"{student_to_update.first_name} {student_to_update.last_name} is now the Assistant Program Officer.")
    else:
        messages.error(request, "No active enrollment found for this student in this program.")
    return redirect('academic_program:student_list', program_name=program_name, program_type=program_type)


def export_students(request, program_name, program_type):
    # Retrieve the program by name and ensure it's active
    program = get_object_or_404(Program, program_name=program_name, flag=True)

    # Retrieve students using the same logic as in your student_list view.
    if program.alumni_program:
        active_enrollments = StudentEnrollment.objects.filter(
            program=program, active=True
        ).exclude(status__in=["Alumnus", "Completed", "Inactive", "Deferred"])
        students = StudentDetail.objects.filter(studentenrollment__in=active_enrollments).distinct()
    else:
        students = ProgramParticipant.objects.filter(program=program)

    # Prepare the filename using the program name and current timestamp
    timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
    filename = f"{program.program_name}-export-{timestamp}.csv"

    # Create the CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)

    # Write CSV header. Include columns conditionally.
    header = ["First Name", "Last Name"]
    if program.alumni_program:
        header += ["Student ID", "Gender"]
    header += ["Company", "Position", "Email", "Phone Number", "Role"]
    writer.writerow(header)

    # Write each student's data, computing the Role based on program officer fields.
    for student in students:
        roles = []
        if program.program_officer and program.program_officer.student == student:
            roles.append("PO")
        if program.assistant_program_officer and program.assistant_program_officer.student == student:
            roles.append("APO")
        role_str = ", ".join(roles) if roles else ""

        row = [student.first_name, student.last_name]
        if program.alumni_program:
            row += [student.unique_id, student.gender]
        row += [student.company, student.position, student.email, student.phone_number, role_str]
        writer.writerow(row)

    return response


def course_details(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    participants = CourseParticipant.objects.filter(course=course)
    students = StudentDetail.objects.filter(studentenrollment__courseparticipant__in=participants)
    lecturer = course.lecturer
    lecturers = Lecturer.objects.all()  # Get all lecturers for the dropdown

    context = {
        'course': course,
        'students': students,
        'lecturer': lecturer,
        'participants': participants,
        'lecturers': lecturers,
    }
    return render(request, 'academic_program/layouts/course_details.html', context)


def courses_under_program_for_attendance(request, program_name):
    program = models.Program.objects.get(program_name=program_name)
    courses_under_program = models.Course.objects.filter(program=program)
    context = {
        'courses': courses_under_program,
        'program': program.program_name
    }
    return render(request, "academic_program/layouts/attendance_courses.html", context=context)


def attendance_overview(request):
    programs = Program.objects.all().order_by('program_name')

    # Get filters from GET parameters.
    selected_program_id = request.GET.get('program')
    selected_date = request.GET.get('date')
    selected_course_id = request.GET.get('course')

    courses = []
    dates = []
    attendance_records_list = []  # to hold the final processed attendance records

    if selected_program_id:
        program = get_object_or_404(Program, id=selected_program_id)
        courses = Course.objects.filter(program=program).order_by('course_name')
        # Get distinct dates for which attendance was recorded.
        dates = Attendance.objects.filter(course__program=program).values_list('attendance_date',
                                                                               flat=True).distinct().order_by(
            '-attendance_date')

        if selected_date and selected_course_id:
            course = get_object_or_404(Course, id=selected_course_id, program=program)
            attendance_records = Attendance.objects.filter(attendance_date=selected_date, course=course)
            participants = CourseParticipant.objects.filter(course=course, flag=True)
            for participant in participants:
                try:
                    # Assuming CourseParticipant.student is a StudentEnrollment
                    enrollment = StudentEnrollment.objects.get(program=program, student=participant.student.student)
                except StudentEnrollment.DoesNotExist:
                    continue
                record = attendance_records.filter(student=enrollment).first()
                attendance_records_list.append({
                    'student_name': f"{participant.student.student.first_name} {participant.student.student.last_name}",
                    'is_present': record.is_present if record else False,
                    'comment': record.comment if record else 'No comment'
                })

    context = {
        'programs': programs,
        'courses': courses,
        'dates': dates,
        'selected_program': selected_program_id,
        'selected_date': selected_date,
        'selected_course': selected_course_id,
        'attendance_records': attendance_records_list,
    }
    return render(request, 'academic_program/layouts/attendance/overview.html', context)


def take_attendance(request, program_name, course_name):
    """
    View to take student attendance for a given program's course,
    with no pre-populated attendance data except linking each form to an enrollment.
    Date is automatically filled to today, but user can change it if they like.
    """
    # 1) Retrieve the Program and Course
    program = get_object_or_404(Program, program_name=program_name, flag=True)
    course = get_object_or_404(Course, course_name=course_name, program=program)

    # 2) Fetch all enrollments for the program (e.g., only active if you prefer).
    enrollments = StudentEnrollment.objects.filter(program=program)
    if not enrollments.exists():
        messages.warning(request, f"No students enrolled in {program_name}.")
        return redirect('academic_program:some_fallback_view')  # Change as appropriate

    if request.method == 'POST':
        # ---------------------------
        # Process the submitted attendance
        # ---------------------------
        date_posted = request.POST.get('date_posted', '').strip()
        if not date_posted:
            messages.error(request, "Please provide a valid date.")
            return redirect('academic_program:take_attendance', program_name=program_name, course_name=course_name)

        try:
            date_obj = datetime.datetime.strptime(date_posted, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid date format. Use YYYY-MM-DD.")
            return redirect('academic_program:take_attendance', program_name=program_name, course_name=course_name)

        error_count = 0
        # Iterate all enrollments and build form instance for each
        for enrollment in enrollments:
            prefix_str = str(enrollment.id)
            form = AttendanceForm(
                data=request.POST,
                student_enrollment=enrollment,
                prefix=prefix_str
            )
            if form.is_valid():
                # Retrieve student_enrollment ID from the hidden field
                student_enrollment_id = form.cleaned_data.get('student')
                try:
                    enrollment_instance = StudentEnrollment.objects.get(pk=student_enrollment_id)
                except StudentEnrollment.DoesNotExist:
                    messages.error(request, f"Enrollment {student_enrollment_id} not found. Skipping.")
                    error_count += 1
                    continue

                # Create or update the attendance record
                Attendance.objects.update_or_create(
                    student=enrollment_instance,
                    course=course,
                    attendance_date=date_obj,
                    defaults={
                        'is_present': form.cleaned_data['is_present'],
                        'comment': form.cleaned_data['comment'],
                        'user': request.user,
                    }
                )
            else:
                # Form not valid for this enrollment
                error_count += 1

        # Provide feedback
        if error_count == 0:
            messages.success(request, "Attendance saved successfully!")
        else:
            messages.warning(request, f"Some entries had errors: {error_count} issue(s).")
        return redirect('academic_program:take_attendance', program_name=program_name, course_name=course_name)

    else:
        # ---------------------------
        # GET: Build blank forms except for linking each form to an enrollment ID
        # and set the date to today in the template.
        # ---------------------------
        attendance_forms = []
        for enrollment in enrollments:
            prefix_str = str(enrollment.id)
            # Provide the student-enrollment ID in hidden field 'student'
            form = AttendanceForm(
                student_enrollment=enrollment,
                prefix=prefix_str,
                initial={'student': enrollment.id}
            )
            attendance_forms.append((enrollment, form))

        today_str = timezone.now().date().strftime('%Y-%m-%d')

        context = {
            'program': program.program_name,
            'course': course.course_name,
            'attendance_forms': attendance_forms,
            'default_date': today_str,  # We'll use this in the template's <input type="date"> value
        }
        return render(request, "academic_program/layouts/attendance_page.html", context)



def export_attendance(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    attendance_records = models.Attendance.objects.filter(course=course).select_related('student')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{course.course_name}_attendance.xlsx"'
    writer = pd.ExcelWriter(response, engine='openpyxl')

    # Group by date and create a worksheet for each date
    for date, group in attendance_records.groupby('attendance_date'):
        df = pd.DataFrame({
            'Student ID': [record.student.student_id for record in group],
            'Student Name': [f"{record.student.first_name} {record.student.last_name}" for record in group],
            'Present': [record.is_present for record in group],
            'Comment': [record.comment for record in group],
        })
        df.to_excel(writer, sheet_name=str(date), index=False)

    writer.save()
    return response


@require_POST
def update_enrollment_status(request, enrollment_id):
    enrollment = get_object_or_404(StudentEnrollment, id=enrollment_id)
    status = request.POST.get('status')

    if status:
        enrollment.status = status
        if status == "Completed":
            enrollment.active = False
        elif status == "Inactive":
            enrollment.active = False
        enrollment.save()
        messages.success(request, "Status Updated")

    return redirect('students:student_details', student_id=enrollment.student.unique_id)


# Program Schedule Views
def select_program_to_add_schedule(request):
    """
    Group all flagged programs by their ProgramCover.
    Within each cover group, separate ongoing from ended programs.
    A program is considered ended if its program_ended flag is True or its end_date is set and is before today.
    """
    programs = Program.objects.filter(flag=True).select_related('program_cover', 'program_type').order_by(
        'program_cover__name', 'program_name')
    grouped_programs = {}  # key: program_cover (or None), value: dict with 'ongoing' and 'ended' lists
    today = timezone.now().date()

    for prog in programs:
        if prog.program_ended or (prog.end_date and prog.end_date < today):
            status = 'ended'
        else:
            status = 'ongoing'
        cover = prog.program_cover  # may be None
        if cover not in grouped_programs:
            grouped_programs[cover] = {'ongoing': [], 'ended': []}
        grouped_programs[cover][status].append(prog)

    context = {
        'grouped_programs': grouped_programs,
    }
    return render(request, "academic_program/layouts/scheduling/add_schedule_overview.html", context)


@login_required(login_url='account:login')
def add_program_schedule(request, program_id):
    program = get_object_or_404(Program, id=program_id, flag=True)
    ProgramScheduleFormSet = modelformset_factory(
        ProgramSchedule,
        form=ProgramScheduleForm,
        extra=1,
        can_delete=True
    )
    schedule_qs = ProgramSchedule.objects.filter(program=program, flag=True)

    if request.method == 'POST':
        schedule_formset = ProgramScheduleFormSet(
            request.POST,
            queryset=schedule_qs,
            form_kwargs={'program': program},
            prefix='schedule'
        )
        if schedule_formset.is_valid():
            for form in schedule_formset:
                if form.cleaned_data:
                    if form.cleaned_data.get('DELETE', False) and form.instance.pk:
                        # Delete the schedule and its associated dates
                        form.instance.delete()
                        continue
                    elif not form.cleaned_data.get('DELETE', False):
                        schedule_obj = form.save(commit=False)
                        schedule_obj.program = program
                        schedule_obj.save()
                        # Process session_dates field
                        dates_str = form.cleaned_data.get('session_dates', '')
                        new_dates = set()
                        if dates_str:
                            # Replace newlines with commas, then split on commas
                            date_list = [d.strip() for d in dates_str.replace('\r\n', ',').replace('\n', ',').split(',')
                                         if d.strip()]
                            for ds in date_list:
                                try:
                                    session_date = datetime.datetime.strptime(ds, '%Y-%m-%d').date()
                                    new_dates.add(session_date)
                                    if not ProgramScheduleDate.objects.filter(
                                        program_schedule=schedule_obj, session_date=session_date
                                    ).exists():
                                        ProgramScheduleDate.objects.create(
                                            program_schedule=schedule_obj,
                                            session_date=session_date
                                        )
                                except ValueError:
                                    messages.warning(request, f"Invalid date format: '{ds}'. Use YYYY-MM-DD.")
                        # Delete any existing session dates that are not in new_dates
                        existing_dates = ProgramScheduleDate.objects.filter(program_schedule=schedule_obj)
                        for existing_date in existing_dates:
                            if existing_date.session_date not in new_dates:
                                existing_date.delete()
            messages.success(request, "Program schedule(s) and dates updated successfully!")
            return redirect('academic_program:add_schedule', program_id=program.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        schedule_formset = ProgramScheduleFormSet(
            queryset=schedule_qs,
            form_kwargs={'program': program},
            prefix='schedule'
        )

    context = {
        'program': program,
        'schedule_formset': schedule_formset,
    }
    return render(request, 'academic_program/layouts/scheduling/add_program_schedule.html', context)






def delete_schedule(request, schedule_id):
    schedule = get_object_or_404(models.ProgramSchedule, id=schedule_id)
    program_id = schedule.course.program.id  # Save the program ID to redirect back to the schedule page
    schedule.flag = False
    schedule.save()
    messages.success(request, "Schedule deleted successfully.")
    return redirect('academic_program:add_schedule', program_id=program_id)


def update_program_schedule(request, schedule_id):
    schedule = get_object_or_404(models.ProgramSchedule, id=schedule_id, flag=True)

    if request.method == 'POST':
        form = forms.ProgramScheduleForm(request.POST, instance=schedule)
        if form.is_valid():
            form.save()
            messages.success(request, "Schedule updated successfully.")
            return redirect('academic_program:add_schedule', program_id=schedule.course.program.id)
    else:
        form = forms.ProgramScheduleForm(instance=schedule)

    return render(request, 'academic_program/layouts/scheduling/update_program_schedule.html', {
        'form': form,
        'program': schedule.course.program,
    })


@login_required
def export_attendance_excel(request, program_id):
    """
    Generate an Excel file with attendance details for all courses under a program.
    Each course gets its own worksheet listing participants’ attendance (P/A) for each distinct date,
    plus summary sections (top 5, lowest 10). An overall summary sheet is included.
    """
    # Retrieve program and its courses
    program = get_object_or_404(Program, id=program_id, flag=True)
    courses = Course.objects.filter(program=program, flag=True)

    # Create workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    overall_summary = []  # To hold summary info per course

    # Process each course
    for course in courses:
        ws = wb.create_sheet(title=course.course_name[:31])  # Sheet names max 31 characters

        # Retrieve distinct attendance dates for this course
        attendance_dates = Attendance.objects.filter(course=course).values_list('attendance_date', flat=True).distinct()
        attendance_dates = sorted(set(attendance_dates))

        # Retrieve participants via CourseParticipant
        participants_qs = CourseParticipant.objects.filter(course=course, flag=True)
        participant_ids = participants_qs.values_list('student__student__id', flat=True)
        # Use StudentDetail to get participant info
        participants = StudentDetail.objects.filter(id__in=participant_ids).distinct()
        participants = list(participants)
        participants.sort(key=lambda s: (s.first_name, s.last_name))

        # Build header row: "Student Name", then attendance dates, then "Total Present"
        header = ["Student Name"]
        header.extend([d.strftime("%Y-%m-%d") for d in attendance_dates])
        header.append("Total Present")
        ws.append(header)
        header_font = Font(bold=True)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="BF0008", end_color="BF0008", fill_type="solid")

        # For each participant, check attendance per date
        participant_attendance = []  # To collect summary counts per participant
        for student in participants:
            row = [f"{student.first_name} {student.last_name}"]
            present_count = 0
            for date in attendance_dates:
                # Check if an attendance record exists with is_present True
                present = Attendance.objects.filter(
                    course=course,
                    attendance_date=date,
                    student__student=student,
                    is_present=True
                ).exists()
                if present:
                    row.append("P")
                    present_count += 1
                else:
                    row.append("A")
            row.append(present_count)
            ws.append(row)
            participant_attendance.append((student, present_count))

        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Add summary section for this course: Top 5 and Lowest 10 attendees
        summary_start = ws.max_row + 2
        ws.cell(row=summary_start, column=1, value="Top 5 Attendees").font = Font(bold=True)
        ws.cell(row=summary_start + 1, column=1, value="Student Name")
        ws.cell(row=summary_start + 1, column=2, value="Total Present")
        # Sort attendance in descending order
        sorted_attendance = sorted(participant_attendance, key=lambda x: x[1], reverse=True)
        for idx, (student, count) in enumerate(sorted_attendance[:5], start=summary_start + 2):
            ws.cell(row=idx, column=1, value=f"{student.first_name} {student.last_name}")
            ws.cell(row=idx, column=2, value=count)

        # Add Lowest 10 attendees
        summary_low_start = ws.max_row + 2
        ws.cell(row=summary_low_start, column=1, value="Lowest 10 Attendees").font = Font(bold=True)
        ws.cell(row=summary_low_start + 1, column=1, value="Student Name")
        ws.cell(row=summary_low_start + 1, column=2, value="Total Present")
        sorted_low = sorted(participant_attendance, key=lambda x: x[1])
        for idx, (student, count) in enumerate(sorted_low[:10], start=summary_low_start + 2):
            ws.cell(row=idx, column=1, value=f"{student.first_name} {student.last_name}")
            ws.cell(row=idx, column=2, value=count)

        # Add overall summary data for this course
        overall_summary.append({
            'course_name': course.course_name,
            'total_participants': len(participants),
            'average_attendance': (sum(x[1] for x in participant_attendance) / len(
                participant_attendance)) if participant_attendance else 0,
        })

    # Create a summary worksheet for overall courses
    ws_overall = wb.create_sheet(title="Course Summary")
    ws_overall.append(["Course Name", "Total Participants", "Average Attendance"])
    for entry in overall_summary:
        ws_overall.append([entry['course_name'], entry['total_participants'], f"{entry['average_attendance']:.2f}"])
    for column_cells in ws_overall.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws_overall.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Prepare HTTP response with Excel file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"Attendance_{program.program_name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_course_attendance_excel(request):
    """
    Exports attendance for a given course and date into an Excel file.
    The file includes:
      - A header row (Student Name, Present, Comment)
      - One row per attendance record (even if absent)
      - A summary section at the end with totals.
    The exported file is named with the course name and date.
    """
    course_id = request.GET.get('course_id')
    date_str = request.GET.get('date')
    if not course_id or not date_str:
        messages.error(request, "Please select a course and a date for export.")
        return HttpResponse("Missing parameters.", status=400)

    try:
        export_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "Invalid date format. Use YYYY-MM-DD.")
        return HttpResponse("Invalid date format.", status=400)

    course = get_object_or_404(Course, id=course_id)
    attendance_records = Attendance.objects.filter(course=course, attendance_date=export_date)

    # Create a new Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Define header row and write it
    headers = ["Student Name", "Present", "Comment"]
    ws.append(headers)

    # Style the header row
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    header_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill

    # Initialize summary counters
    total_participants = attendance_records.count()
    present_count = attendance_records.filter(is_present=True).count()
    absent_count = total_participants - present_count

    # Write attendance records (include those with is_present False)
    for record in attendance_records:
        is_present_text = "Yes" if record.is_present else "No"
        # Construct full student name assuming CourseParticipant.student points to a StudentEnrollment
        student_full_name = f"{record.student.student.first_name} {record.student.student.last_name}"
        row = [student_full_name, is_present_text, record.comment or ""]
        ws.append(row)

    # Add an empty row then summary header and summary values
    ws.append([])
    summary_headers = ["Total Participants", "Present", "Absent"]
    ws.append(summary_headers)
    ws.append([total_participants, present_count, absent_count])

    # Optionally, adjust column widths based on content
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 4

    # Save the workbook to an in-memory output stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Prepare HTTP response with the Excel file
    filename = f"{course.course_name}_attendance_{date_str}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response






