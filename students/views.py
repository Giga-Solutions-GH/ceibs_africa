import csv
import datetime
import io
import ics
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.tokens import default_token_generator
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.db.models import Count, Subquery, OuterRef, Q
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.workbook import Workbook

from academic_program.forms import AttendanceForm
from .tasks import send_alumnus_congratulations_email, send_enrollment_notification_email, send_student_welcome_email, \
    send_enrollment_notification_email_task

from academic_program.models import Course, CourseParticipant, ProgramParticipant, Program, ProgramSchedule, \
    TranscriptRequest, Attendance
from account.models import CustomUser, Role
from alumni.models import Alumni
from finance.models import StudentFinance, PaymentTrail, FinanceStatement, ProgramFees
from student_grading.models import StudentGrade
from students.forms import AddStudentDetailForm, StudentDetailForm, StudentEnrollmentForm, StudentForm, \
    EditStudentContactForm
from students.models import StudentEnrollment, StudentDetail, StudentDocument


# Create your views here.
@transaction.atomic()
def add_student(request):
    if request.method == "POST":
        form = AddStudentDetailForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.cleaned_data
            new_student = StudentDetail.objects.create(
                first_name=data["first_name"],
                last_name=data["last_name"],
                other_names=data["other_names"],
                date_of_birth=data["date_of_birth"],
                gender=data["gender"],
                unique_id=data["unique_id"],
                email=data["email"],
                phone_number=data["phone_number"],
                address=data["address"],
                position=data["position"],
                next_of_kin=data["next_of_kin"],
                image=data.get('image'),
            )

            print(new_student.email)

            # Create the associated user
            new_user = CustomUser.objects.create(
                username=new_student.unique_id,
                email=new_student.email,
                first_name=new_student.first_name,
                last_name=new_student.last_name,
                is_student=True
            )
            new_user.save()

            new_student.user = new_user
            new_student.save()

            # Assign the "Student" role (create if it doesn't exist)
            student_role, _ = Role.objects.get_or_create(name='Student')
            new_user.roles.add(student_role)
            new_user.save()

            # Generate a password reset (set password) token
            uid = urlsafe_base64_encode(force_bytes(new_user.pk))
            token = default_token_generator.make_token(new_user)

            # Queue the welcome email task
            send_student_welcome_email.delay(
                user_id=new_user.pk,
                uidb64=uid,
                token=token
            )

            messages.success(request, "Student added. A welcome email is being sent in the background!")
            return redirect('students:add_student')
        else:
            messages.error(request, "Form is not valid. Please correct the errors.")
    else:
        form = AddStudentDetailForm()

    return render(request, "students/add_student.html", {'form': form})


def upload_students(request):
    if request.method == "POST" and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        df = pd.read_excel(excel_file)
        for _, row in df.iterrows():
            new_student = StudentDetail.objects.create(
                first_name=row['first_name'],
                last_name=row['last_name'],
                other_names=row['other_names'],
                date_of_birth=row['date_of_birth'],
                gender=row['gender'],
                unique_id=row['unique_id'],
                email=row['email'],
                phone_number=row['phone_number'],
                address=row['address'],
                position=row['position'],
                next_of_kin=row['next_of_kin'],
                status=row['status'],
            )
            new_student.save()

            new_user = CustomUser.objects.create(
                username=new_student.unique_id,
                email=new_student.email,
                first_name=new_student.first_name,
                last_name=new_student.last_name,
            )
            new_user.save()
            new_student.user = new_user
            new_student.save()

            # uid = urlsafe_base64_encode(force_bytes(new_user.pk))
            # token = default_token_generator.make_token(new_user)
            # domain = '127.0.0.1:8000'
            # reset_link = f"http://{domain}/reset/{uid}/{token}/"
            # mail_subject = 'Reset your password'
            # message = f'Please use the following link to reset your password: {reset_link}'
            # send_mail(mail_subject, message, 'admin@yourdomain.com', [new_student.email])

        messages.success(request, "Students Uploaded")
        return redirect('students:upload_students')

    return render(request, "students/upload_students.html")


@login_required(login_url='account:login')
def student_detail_view(request, student_id):
    """
    Displays the student details, enrollments, and associated documents.
    Handles updating the student details form.
    """
    # Retrieve the student by unique_id (or change to pk if needed)
    student = get_object_or_404(StudentDetail, unique_id=student_id)

    if request.method == "POST":
        form = StudentDetailForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, "Student information updated successfully.")
            return redirect('students:student_details', student_id=student_id)
        else:
            messages.error(request, "There were errors in the submitted form. Please correct them below.")
    else:
        form = StudentDetailForm(instance=student)

    # Retrieve all enrollments and documents for the student
    enrollments = StudentEnrollment.objects.filter(student=student)
    documents = StudentDocument.objects.filter(student=student)

    context = {
        'form': form,
        'student': student,
        'enrollments': enrollments,
        'documents': documents,
    }
    return render(request, 'students/student_enrollments.html', context)


def delete_student_view(request, student_id):
    student = get_object_or_404(StudentDetail, id=student_id)
    if request.method == "POST":
        student.flag = False
        student.save()
        messages.success(request, "Student has been set to inactive.")
        return redirect('..')  # Update this to your actual student list URL

    context = {
        'student': student,
    }
    messages.info(request, "Student has been deleted.")
    return redirect('students:student_detail', student_id=student_id)


def all_students(request):
    # Retrieve filter parameters from request
    gender = request.GET.get('gender')
    active_program = request.GET.get('active_program')
    company = request.GET.get('company')
    module = request.GET.get('module')
    session = request.GET.get('session')
    name = request.GET.get('name')
    student_number = request.GET.get('student_number')
    email = request.GET.get('email')
    enrollment_status = request.GET.get('status')  # Changed to enrollment_status

    # Filter students based on criteria
    students = StudentDetail.objects.all()

    po_ids = Program.objects.filter(flag=True).values_list('program_officer', flat=True)
    apo_ids = Program.objects.filter(flag=True).values_list('assistant_program_officer', flat=True)

    if gender:
        students = students.filter(gender=gender)
    if active_program:
        students = students.filter(studentenrollment__program__program_name=active_program)
    if company:
        students = students.filter(company__icontains=company)
    if module:
        students = students.filter(studentenrollment__module__icontains=module)
    if session:
        students = students.filter(studentenrollment__session__icontains=session)
    if name:
        students = students.filter(
            Q(first_name__icontains=name) |
            Q(last_name__icontains=name) |
            Q(other_names__icontains=name)
        )
    if student_number:
        students = students.filter(unique_id=student_number)
    if email:
        students = students.filter(email__icontains=email)
    if enrollment_status:  # Use enrollment status for filtering
        students = students.filter(studentenrollment__status=enrollment_status)

    # Annotate with active program and active enrollment status
    active_enrollment = StudentEnrollment.objects.filter(
        student=OuterRef('pk'), active=True
    )

    students = students.annotate(
        active_program=Subquery(active_enrollment.values('program__program_name')[:1]),
        active_status=Subquery(active_enrollment.values('status')[:1])  # Get the active enrollment status
    )

    context = {
        'students': students,
        'genders': ['Male', 'Female'],
        'active_programs': Program.objects.values_list('program_name', flat=True).distinct(),
        'companies': StudentDetail.objects.values_list('company', flat=True).distinct(),
        'modules': StudentEnrollment.objects.values_list('program__course', flat=True).distinct(),
        'statuses': ['Active', 'Inactive', 'Deferred', 'Completed', 'Alumnus'],  # Ensure statuses reflect enrollment
    }
    return render(request, "students/student_program_overview.html", context=context)


def student_list(request, program_name, program_type):
    program = Program.objects.get(program_name=program_name, flag=True)
    if program.alumni_program:
        if request.method == "POST":
            program = Program.objects.get(program_name=program_name, flag=True)
            file = request.FILES['student_file']

            df = pd.read_excel(file)
            counter = 0
            for index, row in df.iterrows():
                student = StudentDetail.objects.create(
                    first_name=row['first name'],
                    last_name=row['last name'],
                    other_names=row['other names'],
                    email=row['email'],
                    date_of_birth=row['date of birth'],
                    gender=row['gender'],
                    unique_id=row['student id'],
                    phone_number=row['phone number'],
                    address=row['address'],
                )
                student.save()

                # Create CustomUser and set password
                user = CustomUser.objects.create_user(
                    username=row['student id'],  # or any unique username
                    email=row['email'],
                    password=get_random_string(12)  # Generate a random password
                )
                user.save()

                new_enrollment = StudentEnrollment.objects.create(
                    student=student,
                    program=program,
                    start_date=datetime.datetime.now(),
                    end_date=program.end_date,
                    active=True
                )
                new_enrollment.save()
                counter += 1
            messages.success(request, f"{counter} Students Enrolled Successfully")

        program = Program.objects.get(program_name=program_name)
        all_students_in_program = StudentDetail.objects.filter(studentenrollment__program=program)
        all_students_in_program = StudentDetail.objects.filter(studentenrollment__program=program)

        # Count enrollment statuses
        enrollment_counts = StudentEnrollment.objects.filter(program=program).values('status').annotate(
            count=Count('status'))
        count = all_students_in_program.count()
        status_counts = {status['status']: status['count'] for status in enrollment_counts}

        # Default counts for statuses that may not exist
        active_count = status_counts.get('Active', 0)
        inactive_count = status_counts.get('Inactive', 0)
        deferred_count = status_counts.get('Deferred', 0)
        completed_count = status_counts.get('Completed', 0)
        alumnus_count = status_counts.get('Alumnus', 0)
        male_count = all_students_in_program.filter(gender='Male').count()
        female_count = all_students_in_program.filter(gender='Female').count()

        context = {
            'students': all_students_in_program,
            'count': count,
            'male_count': male_count,
            'female_count': female_count,
            'program_name': program_name,
            'program': program,
            'active_count': active_count,
            'inactive_count': inactive_count,
            'deferred_count': deferred_count,
            'completed_count': completed_count,
            'alumnus_count': alumnus_count,
        }
        return render(request, "students/student_list.html", context=context)
    else:
        if request.method == "POST":
            program = Program.objects.get(program_name=program_name, flag=True)
            file = request.FILES['student_file']

            df = pd.read_excel(file)
            counter = 0
            for index, row in df.iterrows():
                program_participant = ProgramParticipant.objects.create(
                    first_name=row['first name'],
                    last_name=row['last name'],
                    other_names=row['other names'],
                    email=row['email'],
                    phone_number=row['phone number'],
                    position=row['position'],
                    company=row['company'],
                    program=program
                )

                program_participant.save()

                # Create CustomUser and set password
                user = CustomUser.objects.create_user(
                    username=row['email'],  # or any unique username
                    email=row['email'],
                )
                user.save()
                counter += 1
            messages.success(request, f"{counter} Participants Added Successfully")

        program = Program.objects.get(program_name=program_name)
        all_students_in_program = ProgramParticipant.objects.filter(program=program)
        count = all_students_in_program.count()

        context = {
            'students': all_students_in_program,
            'count': count,
            'program_name': program_name,
            'program': program
        }
        return render(request, "students/student_list.html", context=context)


@transaction.atomic
def enrollment(request):
    if request.method == 'POST':
        form = StudentEnrollmentForm(request.POST)
        if form.is_valid():
            enrollment = form.save()
            # Enroll the student in all courses under the program
            program = enrollment.program
            # In this context, enrollment.student is the active StudentEnrollment instance;
            # however, for linking courses, we need the actual StudentDetail.
            program_fees = ProgramFees.objects.get_or_create(program=program)
            finance_statement = FinanceStatement.objects.get_or_create(program_fees=program_fees)
            new_student = enrollment.student
            new_student_finance = StudentFinance.objects.create(
                student=new_student,
                finance_statement=finance_statement,
            )
            courses = Course.objects.filter(program=program)
            for course in courses:
                CourseParticipant.objects.create(course=course, student=enrollment)
            messages.success(request, "Student Enrolled Successfully and added to all courses in the program.")

            # Prepare links for the email (adjust reverse names as needed)
            from django.urls import reverse
            domain = "ab71-41-215-169-36.ngrok-free.app"  # e.g., "example.com"; consider adding protocol if necessary
            dashboard_link = f"https://{domain}{reverse('students:student_portal')}"
            support_link = ""  # Replace 'support' with your support URL

            # Queue the background task to send the enrollment email
            send_enrollment_notification_email_task.delay(new_student.id, program.id, dashboard_link, support_link)

            return redirect('students:enrol_student')  # Replace with your desired redirect view
    else:
        form = StudentEnrollmentForm()

    return render(request, 'students/enrollment.html', {'form': form})


@transaction.atomic
def enrollment_detail_view(request, enrollment_id):
    """
    Manage course enrollment/unenrollment for a specific StudentEnrollment record.
    """
    # 1) Retrieve the StudentEnrollment
    student_enrollment = get_object_or_404(StudentEnrollment, id=enrollment_id)
    # The actual StudentDetail object
    actual_student = student_enrollment.student
    # Program from that enrollment
    program = student_enrollment.program

    # 2) All courses in the program
    all_courses = Course.objects.filter(program=program)

    # 3) Courses the enrollment is already in, i.e., CourseParticipant with student=that enrollment
    enrolled_courses = Course.objects.filter(
        participants__student=student_enrollment
    )

    # 4) Courses not yet enrolled
    available_courses = all_courses.exclude(id__in=enrolled_courses.values_list('id', flat=True))

    if request.method == 'POST':
        print(request.POST)

        print("got here")
        # Gather the selected course IDs
        selected_course_ids = request.POST.getlist('selected_courses')
        print(selected_course_ids)
        newly_enrolled = []

        for cid in selected_course_ids:
            course = get_object_or_404(Course, id=cid, program=program)
            # 5) If your CourseParticipant references StudentEnrollment:
            #    'student=student_enrollment' is correct
            cp, created = CourseParticipant.objects.get_or_create(
                course=course,
                student=student_enrollment
            )
            if created:
                newly_enrolled.append(course)

        if newly_enrolled:
            messages.success(request, "Enrolled in selected courses successfully.")

            # Optionally send an email with the schedule
            program_schedule_qs = ProgramSchedule.objects.filter(program=program)
            schedule_table_html = render_to_string(
                'students/emails/program_schedule_table.html',
                {'program_schedule': program_schedule_qs}
            )
            # Send Celery email asynchronously
            send_enrollment_notification_email.delay(
                student_email=actual_student.email,
                course_list=[course.course_name for course in newly_enrolled],
                program_name=program.program_name,
                schedule_table_html=schedule_table_html
            )
        else:
            messages.info(request, "No new courses were selected for enrollment.")

        return redirect('students:enrollment_detail', enrollment_id=student_enrollment.id)

    # 6) Get the student's other enrollments (besides this one)
    previous_enrollments = StudentEnrollment.objects.filter(
        student=actual_student
    ).exclude(id=student_enrollment.id)

    context = {
        'enrollment': student_enrollment,
        'program': program,
        'student': actual_student,  # StudentDetail object
        'all_courses': all_courses,
        'enrolled_courses': enrolled_courses,
        'available_courses': available_courses,
        'previous_enrollments': previous_enrollments,
        'is_alumni_program': program.alumni_program,
    }
    return render(request, 'students/enrollment_detail.html', context)


@transaction.atomic
def student_enrollments(request, student_id):
    student = StudentDetail.objects.get(unique_id=student_id)
    enrollments = StudentEnrollment.objects.filter(student=student)
    form = AddStudentDetailForm(
        initial={
            'first_name': student.first_name,
            'last_name': student.last_name,
            'other_names': student.other_names,
            'date_of_birth': student.date_of_birth,
            'gender': student.gender,
            'student_id': student.unique_id,
            'email': student.email,
            'phone_number': student.phone_number,
            'address': student.address
        }
    )
    context = {
        'student': student,
        'enrollments': enrollments,
        'form': form
    }
    return render(request, "students/student_enrollments.html", context=context)


def student_statistics(request):
    # Optional program filter for gender stats
    selected_program_id = request.GET.get('program_id')
    if selected_program_id:
        try:
            selected_program = get_object_or_404(Program, pk=selected_program_id, flag=True)
        except:
            selected_program = None
        if selected_program:
            student_qs = StudentDetail.objects.filter(studentenrollment__program=selected_program).distinct()
        else:
            student_qs = StudentDetail.objects.all()
    else:
        selected_program = None
        student_qs = StudentDetail.objects.all()

    # 1. Gender Distribution Stats
    gender_stats = student_qs.values('gender').annotate(count=Count('id')).order_by('gender')

    # 2. Students per Program Stats (all programs)
    program_stats = (
        StudentEnrollment.objects.values('program__program_name')
        .annotate(count=Count('id'))
        .order_by('program__program_name')
    )

    # 3. Students per Course Stats (all courses)
    from academic_program.models import Course  # Ensure Course is imported
    course_stats = (
        Course.objects.annotate(student_count=Count('participants'))
        .values('course_name', 'student_count')
        .order_by('course_name')
    )

    # 4. Students per Year Stats
    year_stats = (
        StudentEnrollment.objects.values('start_date__year')
        .annotate(count=Count('id'))
        .order_by('start_date__year')
    )

    # 5. Students by Company
    company_stats = (
        StudentDetail.objects.values('company')
        .exclude(company__isnull=True).exclude(company__exact="")
        .annotate(count=Count('id'))
        .order_by('company')
    )

    # 6. Students by Position
    position_stats = (
        StudentDetail.objects.values('position')
        .exclude(position__isnull=True).exclude(position__exact="")
        .annotate(count=Count('id'))
        .order_by('position')
    )

    # 7. Students by Nationality
    nationality_stats = (
        StudentDetail.objects.values('nationality')
        .exclude(nationality__isnull=True).exclude(nationality__exact="")
        .annotate(count=Count('id'))
        .order_by('nationality')
    )

    # For the program filter dropdown, list all active programs
    all_programs = Program.objects.filter(flag=True).order_by('program_name')

    context = {
        "selected_program": selected_program,
        "all_programs": all_programs,
        "gender_stats": gender_stats,
        "program_stats": program_stats,
        "course_stats": course_stats,
        "year_stats": year_stats,
        "company_stats": company_stats,
        "position_stats": position_stats,
        "nationality_stats": nationality_stats,
    }
    return render(request, "students/student_statistics.html", context)


def export_student_statistics(request):
    # Gather data
    gender_stats = StudentDetail.objects.values('gender').annotate(count=Count('id')).order_by('gender')
    program_stats = StudentEnrollment.objects.values('program__program_name').annotate(
        count=Count('id')).order_by('program__program_name')
    course_stats = Course.objects.annotate(student_count=Count('participants')).values('course_name',
                                                                                       'student_count').order_by(
        'course_name')
    year_stats = StudentEnrollment.objects.values('start_date__year').annotate(count=Count('id')).order_by(
        'start_date__year')

    # Create a new Excel workbook and sheets
    wb = Workbook()

    # Gender Distribution
    ws = wb.active
    ws.title = "Gender Distribution"
    ws.append(["Gender", "Count"])
    for item in gender_stats:
        ws.append([item['gender'], item['count']])

    # Students per Program
    ws = wb.create_sheet(title="Students per Program")
    ws.append(["Program", "Count"])
    for item in program_stats:
        ws.append([item['program__program_name'], item['count']])

    # Students per Course
    ws = wb.create_sheet(title="Students per Course")
    ws.append(["Course", "Number of Students"])
    for item in course_stats:
        ws.append([item['course_name'], item['student_count']])

    # Students per Year
    ws = wb.create_sheet(title="Students per Year")
    ws.append(["Year", "Count"])
    for item in year_stats:
        ws.append([item['start_date__year'], item['count']])

    # Save workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create an HTTP response with the Excel file
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=student_statistics.xlsx'
    return response


@transaction.atomic
def add_student_from_program(request, program_id):
    program = get_object_or_404(Program, id=program_id)

    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            # Create the student detail
            student = form.save()

            # Create the associated user
            user = CustomUser.objects.create_user(
                username=student.unique_id,
                email=student.email,
                password=student.unique_id,  # Change this as needed
                first_name=student.first_name,
                last_name=student.last_name
            )
            student.user = user
            student.save()

            # Enroll student in the program
            StudentEnrollment.objects.create(
                student=student,
                program=program,
                start_date=program.start_date,
                end_date=program.end_date,
                status="Active"
            )

            # Enroll student in all courses for this program
            courses = Course.objects.filter(program=program)
            for course in courses:
                CourseParticipant.objects.create(
                    course=course,
                    student=student
                )

            return redirect('students:add_student_from_program_page',
                            program_id=program_id)  # Redirect to student list page

    else:
        form = StudentForm()

    context = {
        'form': form,
        'program': program,
    }
    return render(request, 'students/add_student_from_program_page.html', context)


##################################################################################################################################################################################################################################################
# Student Portal
##################################################################################################################################################################################################################################################


@login_required(login_url='account:login')
def student_portal(request):
    try:
        student = get_object_or_404(StudentDetail, user=request.user)
    except ObjectDoesNotExist:
        messages.error(request, "Student record not found.")
        return redirect('account:login')

    # Get student enrollments (if none, use an empty list)
    student_enrollments = StudentEnrollment.objects.filter(student=student)
    # Try to get an active enrollment; if none, set active_program to None.
    try:
        active_enrollment = StudentEnrollment.objects.get(student=student, active=True)
        active_program = active_enrollment.program
    except StudentEnrollment.DoesNotExist:
        active_program = None

    # Build enrolled programs details (program + list of courses)
    enrolled_programs = []
    for enrollment in student_enrollments:
        program = enrollment.program
        courses = Course.objects.filter(program=program)
        enrolled_programs.append({
            'program': program,
            'courses': courses,
            'enrollment': enrollment,
        })

    # Get student's finance records
    student_finance = StudentFinance.objects.filter(student=student)

    # Get student's documents
    student_documents = StudentDocument.objects.filter(student=student)

    context = {
        'student': student,
        'active_program': active_program,
        'student_enrollments': student_enrollments,
        'enrolled_programs': enrolled_programs,
        'student_finance': student_finance,
        'student_documents': student_documents,
    }
    return render(request, 'students/student_portal/student_portal.html', context)


@login_required(login_url='account:login')
def edit_student_contact(request):
    if request.method == 'POST':
        form = EditStudentContactForm(request.POST)
        if form.is_valid():
            student = get_object_or_404(StudentDetail, user=request.user)
            student.phone_number = form.cleaned_data['phone_number']
            student.email = form.cleaned_data['email']
            student.company = form.cleaned_data['company']
            student.position = form.cleaned_data['position']
            student.save()
            messages.success(request, 'Contact information updated successfully.')
            return redirect('students:student_portal')
    else:
        student = get_object_or_404(StudentDetail, user=request.user)
        form = EditStudentContactForm(initial={
            'phone_number': student.phone_number,
            'email': student.email,
            'company': student.company,
            'position': student.position
        })

    return render(request, 'students/student_portal/edit_student_contact.html', {'form': form})


@transaction.atomic
@login_required(login_url='account:login')
def request_transcript(request):
    if request.method == 'POST':
        completion_year = request.POST.get('completion_year')
        enrollment_id = request.POST.get('enrollment_program')
        enrollment = get_object_or_404(StudentEnrollment, id=enrollment_id)
        student = get_object_or_404(StudentDetail, user=request.user)

        if TranscriptRequest.objects.filter(program=enrollment.program, student=student,
                                            year_of_completion=completion_year).exists():
            messages.info(request, 'Enrollment request already submitted.')
            return redirect('students:student_portal')

        new_transcript_request = TranscriptRequest.objects.create(
            student=student,
            user=request.user,
            year_of_completion=completion_year,
            program=enrollment.program,
        )
        new_transcript_request.save()

        # Handle transcript request logic here (e.g., sending an email, creating a request record)
        messages.success(request,
                         f'Transcript request for {enrollment.program.program_name} in {completion_year} has been received.')
        return redirect('students:student_portal')
    return redirect('students:student_portal')


@login_required(login_url='account:login')
def student_programs(request):
    user = CustomUser.objects.get(id=request.user.id)
    student = StudentDetail.objects.get(user=user)  # Assuming the user is logged in as a student
    enrollments = StudentEnrollment.objects.filter(student=student)

    context = {
        'enrollments': enrollments
    }
    return render(request, 'students/student_portal/student_programs.html', context)


@login_required(login_url='account:login')
def program_courses(request, enrollment_id):
    enrollment = get_object_or_404(StudentEnrollment, id=enrollment_id)
    courses = Course.objects.filter(program=enrollment.program)
    grades = StudentGrade.objects.filter(student=enrollment.student, program=enrollment.program)

    # Create a list of tuples (course, grade)
    courses_with_grades = []
    for course in courses:
        # Find the grade for this course, if it exists
        grade = next((g for g in grades if g.course_id == course.id), None)
        courses_with_grades.append((course, grade))

    context = {
        'enrollment': enrollment,
        'courses_with_grades': courses_with_grades,
    }
    return render(request, 'students/student_portal/program_courses.html', context)


@login_required(login_url='account:login')
def program_schedule_overview(request):
    user = CustomUser.objects.get(id=request.user.id)
    student = StudentDetail.objects.get(user=user)
    student_programs = StudentEnrollment.objects.filter(active=True, student=student)
    context = {
        "student_programs": student_programs
    }
    return render(request, "students/student_portal/student_programs_for_schedule.html", context=context)


@login_required(login_url='account:login')
def student_program_schedule(request, program_id):
    # Retrieve the program; if not found, redirect with a message.
    program = get_object_or_404(Program, id=program_id, flag=True)

    # Retrieve all schedules for the program and prefetch their session dates.
    program_schedule = ProgramSchedule.objects.filter(program=program, flag=True).prefetch_related('schedule_dates')

    context = {
        'program': program,
        'program_schedule': program_schedule,
        'today': timezone.now().date(),  # in case you need it in template
    }
    return render(request, "students/student_portal/program_schedule.html", context)


@login_required(login_url='account:login')
def transcript_requests(request):
    """
    List transcript requests with server-side filtering by:
     - Generated status (all, generated, not generated)
     - Year of completion
     - Program
    """
    # Get GET parameters (default values provided)
    filter_status = request.GET.get('filter_status', 'all').strip()  # values: 'all', 'true', 'false'
    filter_year = request.GET.get('filter_year', '').strip()
    filter_program = request.GET.get('filter_program', '').strip()

    # Build the initial queryset with related transcript requests prefetched
    qs = TranscriptRequest.objects.prefetch_related('generatedtranscriptrequest_set').all()

    # Apply status filter
    if filter_status != 'all':
        if filter_status == 'true':
            qs = qs.filter(generated=True)
        elif filter_status == 'false':
            qs = qs.filter(generated=False)

    # Apply year filter (exact match, you can use __icontains if needed)
    if filter_year:
        qs = qs.filter(year_of_completion__iexact=filter_year)

    # Apply program filter (by program ID)
    if filter_program:
        qs = qs.filter(program__id=filter_program)

    # Prepare data for dropdowns
    # Get distinct year values from transcript requests
    distinct_years = TranscriptRequest.objects.exclude(year_of_completion='') \
        .values_list('year_of_completion', flat=True).distinct()
    # Get all active programs
    programs = Program.objects.filter(flag=True)

    context = {
        'requests': qs,
        'filter_status': filter_status,
        'filter_year': filter_year,
        'filter_program': filter_program,
        'distinct_years': distinct_years,
        'programs': programs,
    }
    return render(request, 'students/student_portal/transcript_requests.html', context)


@transaction.atomic
def mark_enrollment_as_alumni(request, enrollment_id):
    enrollment = get_object_or_404(StudentEnrollment, id=enrollment_id)
    student = enrollment.student

    # Check if the enrollment is already alumnus
    if enrollment.status == 'Alumnus':
        messages.warning(
            request,
            f"The enrollment for {student.first_name} {student.last_name} is already marked as alumnus."
        )
    else:
        # Mark the enrollment as alumnus
        enrollment.status = 'Alumnus'
        enrollment.active = False
        enrollment.save()

        # Create a record in the Alumni model
        alumnus = Alumni.objects.create(
            student=student,
            user=student.user,
            first_name=student.first_name,
            last_name=student.last_name,
            email=student.email,
            program=enrollment.program,
            year_of_completion=enrollment.end_date.year if enrollment.end_date else 'N/A',
            current_position=student.position,
            company=student.company,
            industry='Tech',  # or any default value you want
            picture=student.image,
        )

        # Trigger the Celery task to send the alumnus confirmation email
        if student.email:
            send_alumnus_congratulations_email.delay(
                student_email=student.email,
                program_name=enrollment.program.program_name,
                year_of_completion=alumnus.year_of_completion
            )

        messages.success(
            request,
            f"{student.first_name} {student.last_name} has been successfully marked as alumnus and added to alumni records."
        )

    return redirect('students:student_overview')


def generate_transcript(request, transcript_id):
    transcript_request = get_object_or_404(TranscriptRequest, id=transcript_id)

    # Get the student's grades for the program
    student_grades = StudentGrade.objects.filter(
        student=transcript_request.student,
        program=transcript_request.program
    )

    # Create the PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="transcript_{transcript_request.id}.pdf"'

    # Create a PDF object using ReportLab
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Add CEIBS logo (Make sure the logo path is correct)
    logo_path = os.path.join(settings.STATIC_ROOT, 'assets/images/logo.png')
    p.drawImage(logo_path, 50, height - 100, width=100, height=50)

    # Title
    p.setFont("Helvetica-Bold", 18)
    p.drawString(200, height - 70, "Transcript of Owner Director Programme 2022")

    # Student details
    p.setFont("Helvetica", 12)
    p.drawString(50, height - 120, f"Student ID: {transcript_request.student.unique_id}")
    p.drawString(250, height - 120,
                 f"Name: {transcript_request.student.first_name} {transcript_request.student.last_name}")
    p.drawString(50, height - 140, f"Academic Year: April 2022 - September 2022")

    # Table header
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, height - 180, "Course")
    p.drawString(400, height - 180, "Grade")
    p.line(50, height - 185, 550, height - 185)

    # Add courses and grades in a table-like format
    y_position = height - 210
    p.setFont("Helvetica", 12)
    for grade in student_grades:
        course_name = grade.course.course_name
        grade_value = grade.grade.grade
        p.drawString(50, y_position, course_name)
        p.drawString(400, y_position, grade_value)
        y_position -= 20  # Move down for the next course

    # Grading system explanation
    p.line(50, y_position - 10, 550, y_position - 10)  # Add a line above the grading explanation
    y_position -= 30
    p.setFont("Helvetica", 10)
    p.drawString(50, y_position,
                 "To meet the requirement for graduation, the student shall successfully complete all required courses,")
    p.drawString(50, y_position - 15, "obtain the required grade and pass the non-credit courses.")
    p.drawString(50, y_position - 35, "Distinction (80-100%), Credit (65-79%), Pass (50-64%), Fail (0-49%)")

    # Footer with issue date and signature
    p.setFont("Helvetica", 12)
    y_position -= 80
    p.drawString(50, y_position, f"Date of Issue: September 10, 2022")
    p.drawString(50, y_position - 15, "CEIBS (Not Valid without Official Seal)")

    # Finalize the PDF
    p.showPage()
    p.save()

    return response


def add_to_calendar(request, schedule_id):
    schedule = get_object_or_404(ProgramSchedule, id=schedule_id, flag=True)

    # Create an ICS event using the ics package (pip install ics)
    event = ics.Event()
    event.name = schedule.course.course_name
    event.begin = datetime.datetime.combine(schedule.start_date, schedule.start_time)
    event.end = datetime.datetime.combine(schedule.end_date, schedule.end_time)
    event.location = schedule.location
    event.description = f"Program: {schedule.program.program_name}"

    calendar = ics.Calendar()
    calendar.events.add(event)

    response = HttpResponse(str(calendar), content_type='text/calendar')
    filename = f"{schedule.program.program_name}_{schedule.course.course_name}_schedule.ics"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url='account:login')
def student_fee_history(request):
    # Retrieve the student based on the current logged-in user.
    student = get_object_or_404(StudentDetail, user=request.user)

    # Retrieve all finance records for the student.
    student_finances = StudentFinance.objects.filter(student=student).select_related(
        'fees__program_fees__program'
    )

    # Build a list of finance details for display.
    finance_data = []
    for sf in student_finances:
        # Retrieve total fee from ProgramFees model
        total_fee = sf.fees.program_fees.fee  # e.g., 5000.0
        fees_paid = sf.fees_paid
        amount_due = total_fee - fees_paid
        finance_data.append({
            'program': sf.fees.program_fees.program.program_name,
            'total_fee': total_fee,
            'fees_paid': fees_paid,
            'student_balance': sf.student_balance,
            'percentage_cleared': sf.percentage_cleared,
            'amount_due': amount_due,
            'finance_id': sf.id,
        })

    # Retrieve all payment trail records for this student, ordered by most recent first.
    payment_trails = PaymentTrail.objects.filter(student_finance__student=student).order_by('-timestamp')

    context = {
        'student': student,
        'finance_data': finance_data,
        'payment_trails': payment_trails,
    }
    return render(request, 'students/student_portal/student_finance.html', context)


@login_required(login_url='account:login')
def program_officer_attendance(request):
    """
    Show the programs for which the logged-in student is assigned as Program Officer (PO)
    or Assistant Program Officer (APO). If only one program exists, redirect to its courses view.
    """
    try:
        student = StudentDetail.objects.get(user=request.user)
    except StudentDetail.DoesNotExist:
        messages.error(request, "Student details not found.")
        return redirect('account:login')

    # Find programs where the student's enrollment is designated as PO or APO.
    programs = Program.objects.filter(
        Q(program_officer__student=student) | Q(assistant_program_officer__student=student)
    ).distinct()

    if not programs.exists():
        messages.info(request, "You are not assigned as a Program Officer.")
        return redirect('students:student_portal')

    if programs.count() == 1:
        program = programs.first()
        return redirect('program_officer_courses', program_id=program.id)

    context = {
        'programs': programs,
    }
    return render(request, 'students/student_portal/program_officer_attendance.html', context)


@login_required(login_url='account:login')
def program_officer_courses(request, program_id):
    """
    Lists courses for the selected program. The program officer can select a course to take attendance.
    """
    program = get_object_or_404(Program, id=program_id, flag=True)
    courses = Course.objects.filter(program=program, flag=True)
    context = {
        'program': program,
        'courses': courses,
    }
    return render(request, 'students/student_portal/program_officer_courses.html', context)


@login_required(login_url='account:login')
def take_attendance_for_course(request, program_id, course_id):
    """
    Allows the program officer to take attendance for a specific course.
    The page shows a date selector (defaulting to today's date) and a form per student enrollment.
    """
    program = get_object_or_404(Program, id=program_id, flag=True)
    course = get_object_or_404(Course, id=course_id, program=program, flag=True)

    # Get all enrollments for the program (adjust logic as needed, e.g., only active enrollments)
    enrollments = StudentEnrollment.objects.filter(program=program)
    if not enrollments.exists():
        messages.warning(request, f"No students are enrolled in {program.program_name}.")
        return redirect('program_officer_courses', program_id=program.id)

    if request.method == 'POST':
        date_posted = request.POST.get('date_posted', '').strip()
        if not date_posted:
            messages.error(request, "Please provide a valid date.")
            return redirect('take_attendance_for_course', program_id=program.id, course_id=course.id)
        try:
            date_obj = datetime.datetime.strptime(date_posted, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
            return redirect('take_attendance_for_course', program_id=program.id, course_id=course.id)

        error_count = 0
        for enrollment in enrollments:
            form = AttendanceForm(request.POST, student_enrollment=enrollment, prefix=str(enrollment.id))
            if form.is_valid():
                student_id = form.cleaned_data.get('student')
                try:
                    enrollment_instance = StudentEnrollment.objects.get(pk=student_id)
                except StudentEnrollment.DoesNotExist:
                    messages.error(request, f"Enrollment ID {student_id} not found. Skipping record.")
                    error_count += 1
                    continue
                Attendance.objects.update_or_create(
                    student=enrollment_instance,
                    course=course,
                    attendance_date=date_obj,
                    defaults={
                        'is_present': form.cleaned_data['is_present'],
                        'comment': form.cleaned_data['comment'],
                        'user': request.user,
                    }
                )
            else:
                error_count += 1

        if error_count == 0:
            messages.success(request, "Attendance records saved successfully!")
        else:
            messages.warning(request, f"{error_count} record(s) had errors. Please review.")
        return redirect('take_attendance_for_course', program_id=program.id, course_id=course.id)
    else:
        # For GET, prepare forms for each enrollment.
        attendance_forms = []
        today = timezone.now().date()
        for enrollment in enrollments:
            existing = Attendance.objects.filter(student=enrollment, course=course, attendance_date=today).first()
            if existing:
                form = AttendanceForm(student_enrollment=enrollment, prefix=str(enrollment.id), initial={
                    'student': enrollment.id,
                    'is_present': existing.is_present,
                    'comment': existing.comment,
                })
            else:
                form = AttendanceForm(student_enrollment=enrollment, prefix=str(enrollment.id), initial={
                    'student': enrollment.id
                })
            attendance_forms.append((enrollment, form))

        context = {
            'program': program,
            'course': course,
            'attendance_forms': attendance_forms,
            'today': today,
        }
        return render(request, 'students/student_portal/take_attendance_for_course.html', context)


@login_required
def export_student_statistics(request):
    """
    Exports student statistics and detailed data to CSV.
    The export includes summary sections grouped by gender, program, course,
    nationality, and company. A checkbox ("include_alumni") allows the user
    to decide if ended programs (alumni) should be included.
    """
    include_alumni = request.GET.get('include_alumni', 'off') == 'on'

    # Get students (if not including alumni, filter by active enrollments)
    if include_alumni:
        students = StudentDetail.objects.all()
    else:
        # Filter students having at least one active enrollment.
        active_student_ids = StudentEnrollment.objects.filter(active=True).values_list('student__id', flat=True)
        students = StudentDetail.objects.filter(id__in=active_student_ids)

    # 1. Summary by Gender
    gender_summary = students.values('gender').annotate(count=Count('id')).order_by('gender')

    # 2. Summary by Program
    program_summary = StudentEnrollment.objects.filter(student__in=students) \
        .values('program__program_name') \
        .annotate(count=Count('id')) \
        .order_by('program__program_name')

    # 3. Summary by Course (via CourseParticipant)
    from academic_program.models import CourseParticipant
    course_summary = CourseParticipant.objects.filter(student__student__in=students) \
        .values('course__course_name') \
        .annotate(count=Count('id')) \
        .order_by('course__course_name')

    # 4. Summary by Nationality (students)
    nationality_summary = students.values('nationality').annotate(count=Count('id')).order_by('nationality')

    # 5. Summary by Company (students)
    company_summary = students.values('company').exclude(company__isnull=True).exclude(company__exact="") \
        .annotate(count=Count('id')).order_by('company')

    # Build CSV response
    response = HttpResponse(content_type='text/csv')
    filename = "Student_Statistics"
    filename += "_with_alumni" if include_alumni else "_active_only"
    filename += f"_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)

    # Write Summary by Gender
    writer.writerow(["Summary by Gender"])
    writer.writerow(["Gender", "Count"])
    for entry in gender_summary:
        writer.writerow([entry['gender'], entry['count']])
    writer.writerow([])

    # Write Summary by Program
    writer.writerow(["Summary by Program"])
    writer.writerow(["Program", "Count"])
    for entry in program_summary:
        writer.writerow([entry['program__program_name'], entry['count']])
    writer.writerow([])

    # Write Summary by Course
    writer.writerow(["Summary by Course"])
    writer.writerow(["Course", "Count"])
    for entry in course_summary:
        writer.writerow([entry['course__course_name'], entry['count']])
    writer.writerow([])

    # Write Summary by Nationality
    writer.writerow(["Summary by Nationality"])
    writer.writerow(["Nationality", "Count"])
    for entry in nationality_summary:
        writer.writerow([entry['nationality'], entry['count']])
    writer.writerow([])

    # Write Summary by Company
    writer.writerow(["Summary by Company"])
    writer.writerow(["Company", "Count"])
    for entry in company_summary:
        writer.writerow([entry['company'], entry['count']])
    writer.writerow([])

    # Write Detailed Student Data
    writer.writerow(["Detailed Student Data"])
    writer.writerow(
        ["Student ID", "First Name", "Last Name", "Gender", "Nationality", "Company", "Email", "Phone Number"])
    for student in students:
        writer.writerow([
            student.unique_id,
            student.first_name,
            student.last_name,
            student.gender,
            student.nationality or "",
            student.company or "",
            student.email,
            student.phone_number
        ])

    return response
